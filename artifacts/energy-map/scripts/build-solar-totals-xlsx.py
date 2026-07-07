"""Generate MCFC_Solar_Totals.xlsx onto the Desktop.

Two sheets:
  1. Solar by Phase - per-array breakdown with phase subtotals (SUM formulas)
     and a grand total (SUM of subtotals).
  2. Annualised comparison - apples-to-apples 12-month view: Phase 1 actuals
     scaled by 0.75 (16-month to 12-month), Phase 2 and 3 as stored.

No em or en dashes anywhere. Uses Arial throughout.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUT = Path(r"C:\Users\james.evans\OneDrive - Clearvolt\Desktop\MCFC_Solar_Totals.xlsx")

# Per-array figures as stored in the map's data panel.
PHASE1 = [
    ("Joie Stadium",        995296),
    ("Performance Centre",  612426),
    ("FM Building",          93780),
    ("TV Studio",            17564),
]
PHASE2 = [
    ("MCWFC Solar",          49733),
    ("Ground Mount 2A",     284628),
    ("Ground Mount 2B",     696635),
]
PHASE3 = [
    ("North Stand Hotel Solar",       84278),
    ("North Stand Commercial Solar",  64595),
    ("Towers Solar",                 119890),
    ("Co-op Live Solar",            1250000),
]

# Styling constants.
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")  # dark blue
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SUBTOTAL_FILL = PatternFill("solid", fgColor="E8EEF7")
SUBTOTAL_FONT = Font(name=FONT_NAME, bold=True, size=11)
GRAND_FILL = PatternFill("solid", fgColor="D6DEF0")
GRAND_FONT = Font(name=FONT_NAME, bold=True, size=11)
BAND_FILL = PatternFill("solid", fgColor="F5F7FA")
BODY_FONT = Font(name=FONT_NAME, size=11)
THIN = Side(style="thin", color="C9CED6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

KWH_FORMAT = "#,##0"

wb = Workbook()

# ── Sheet 1 ─────────────────────────────────────────────────────────────────
s1 = wb.active
s1.title = "Solar by Phase"

headers1 = ["Phase", "Array", "Annual generation (kWh)", "Data type", "Period"]
for col, h in enumerate(headers1, 1):
    c = s1.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER

row = 2

def write_phase(label, items, data_type, period, banding_offset):
    """Write a phase block and return (first_data_row, subtotal_row)."""
    global row
    first = row
    for i, (name, kwh) in enumerate(items):
        s1.cell(row=row, column=1, value=label).font = BODY_FONT
        s1.cell(row=row, column=2, value=name).font = BODY_FONT
        kw = s1.cell(row=row, column=3, value=kwh)
        kw.font = BODY_FONT
        kw.number_format = KWH_FORMAT
        s1.cell(row=row, column=4, value=data_type).font = BODY_FONT
        s1.cell(row=row, column=5, value=period).font = BODY_FONT
        # Banding: alternate rows get light grey fill, based on a counter that
        # ignores the subtotal rows so banding stays consistent.
        if (banding_offset + i) % 2 == 1:
            for col in range(1, 6):
                s1.cell(row=row, column=col).fill = BAND_FILL
        for col in range(1, 6):
            s1.cell(row=row, column=col).border = BORDER
        row += 1
    last = row - 1
    # Subtotal row with SUM formula.
    s1.cell(row=row, column=1, value=label).font = SUBTOTAL_FONT
    s1.cell(row=row, column=2, value=f"{label} subtotal").font = SUBTOTAL_FONT
    sub = s1.cell(row=row, column=3, value=f"=SUM(C{first}:C{last})")
    sub.font = SUBTOTAL_FONT
    sub.number_format = KWH_FORMAT
    s1.cell(row=row, column=4, value="").font = SUBTOTAL_FONT
    s1.cell(row=row, column=5, value="").font = SUBTOTAL_FONT
    for col in range(1, 6):
        s1.cell(row=row, column=col).fill = SUBTOTAL_FILL
        s1.cell(row=row, column=col).border = BORDER
    subtotal_row = row
    row += 1
    return first, last, subtotal_row

p1_first, p1_last, p1_sub = write_phase("Phase 1", PHASE1, "Actual", "16-month (Jan 2025 to Apr 2026)", 0)
p2_first, p2_last, p2_sub = write_phase("Phase 2", PHASE2, "Modelled", "12-month", 0)
p3_first, p3_last, p3_sub = write_phase("Phase 3", PHASE3, "Modelled", "12-month", 0)

# Grand total: SUM of the three subtotal cells.
s1.cell(row=row, column=1, value="").font = GRAND_FONT
s1.cell(row=row, column=2, value="Grand total (all phases, as stored)").font = GRAND_FONT
g = s1.cell(row=row, column=3, value=f"=C{p1_sub}+C{p2_sub}+C{p3_sub}")
g.font = GRAND_FONT
g.number_format = KWH_FORMAT
s1.cell(row=row, column=4, value="").font = GRAND_FONT
s1.cell(row=row, column=5, value="").font = GRAND_FONT
for col in range(1, 6):
    s1.cell(row=row, column=col).fill = GRAND_FILL
    s1.cell(row=row, column=col).border = BORDER

# Column widths.
widths1 = {"A": 10, "B": 36, "C": 24, "D": 12, "E": 34}
for col, w in widths1.items():
    s1.column_dimensions[col].width = w

# Freeze the header row.
s1.freeze_panes = "A2"

# ── Sheet 2 ─────────────────────────────────────────────────────────────────
s2 = wb.create_sheet("Annualised comparison")

headers2 = ["Phase", "Description", "Annual generation (kWh)", "Note"]
for col, h in enumerate(headers2, 1):
    c = s2.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER

rows2 = [
    ("Phase 1", "Existing CFA rooftop (actuals annualised x 0.75)",
     f"='Solar by Phase'!C{p1_sub}*0.75",
     "16-month actuals scaled by 12/16 = 0.75"),
    ("Phase 2", "CFA new arrays",
     f"='Solar by Phase'!C{p2_sub}",
     "as stored (modelled 12-month)"),
    ("Phase 3", "Etihad campus arrays",
     f"='Solar by Phase'!C{p3_sub}",
     "as stored (modelled 12-month)"),
]
r = 2
for i, (phase, desc, formula, note) in enumerate(rows2):
    s2.cell(row=r, column=1, value=phase).font = BODY_FONT
    s2.cell(row=r, column=2, value=desc).font = BODY_FONT
    val = s2.cell(row=r, column=3, value=formula)
    val.font = BODY_FONT
    val.number_format = KWH_FORMAT
    s2.cell(row=r, column=4, value=note).font = BODY_FONT
    if i % 2 == 1:
        for col in range(1, 5):
            s2.cell(row=r, column=col).fill = BAND_FILL
    for col in range(1, 5):
        s2.cell(row=r, column=col).border = BORDER
    r += 1

# Grand total annualised: SUM of the three rows above.
s2.cell(row=r, column=1, value="").font = GRAND_FONT
s2.cell(row=r, column=2, value="Grand total annualised (apples-to-apples)").font = GRAND_FONT
grand = s2.cell(row=r, column=3, value=f"=SUM(C2:C{r-1})")
grand.font = GRAND_FONT
grand.number_format = KWH_FORMAT
s2.cell(row=r, column=4, value="").font = GRAND_FONT
for col in range(1, 5):
    s2.cell(row=r, column=col).fill = GRAND_FILL
    s2.cell(row=r, column=col).border = BORDER

widths2 = {"A": 10, "B": 48, "C": 24, "D": 38}
for col, w in widths2.items():
    s2.column_dimensions[col].width = w
s2.freeze_panes = "A2"

# Default font for any later-added cells.
for sh in wb.worksheets:
    sh.sheet_view.showGridLines = False

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Size: {OUT.stat().st_size:,} bytes")
print(f"Sheets: {wb.sheetnames}")
